from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from zss import simple_distance, Node, distance
import textdistance
from functools import cmp_to_key

import os
import argparse
import json
from glob import glob

from rndtable.utils import load_json, cal_spec_location_iou, combine_text_list, stand2spec_output
from rndtable.visualize import visualize_bar_chart, dump_report_excel

class WeirdNode(Node):
    def __init__(self, label, rowspan, colspan):
        super(WeirdNode, self).__init__(label, children=None)
        self.rowspan = rowspan
        self.colspan = colspan

    @staticmethod
    def get_children(node):
        return node.children

    def get_label(self):
        return self.label
    
    @staticmethod
    def label_dist(label_1, label_2):
        return 1. - textdistance.levenshtein.normalized_similarity(label_1, label_2)
    
    @staticmethod
    def insert_cost(node):
        if len(WeirdNode.get_children(node)) > 0:
            return 0
        return 1

    @staticmethod
    def remove_cost(node):
        if len(WeirdNode.get_children(node)) > 0:
            return 0
        return 1
    
    @staticmethod
    def update_cost(a, b):
        if len(WeirdNode.get_children(a)) > 0 or len(WeirdNode.get_children(b)) > 0:
            return 1
        else:
            if a.rowspan != b.rowspan or a.colspan != b.colspan:
                return 1
            else:
                return WeirdNode.label_dist(a.get_label(), b.get_label()) 
    
    def addkid(self, node, before=False):
        if before:  
            self.children.insert(0, node)
        else: 
            self.children.append(node)
        return self

def convert_sheet2tree(sheet, ocr_included=True):
    def set_node_label(label):
        if ocr_included:
            return label
        return sheet.title
        
    conv = lambda i : i or ''
    tree = WeirdNode(set_node_label(sheet.title), None, None)
    merged_ranges = sheet.merged_cells.ranges.copy()
    for row in range(0, sheet.max_row):
        node = WeirdNode(set_node_label('row_' + str(row+1)), None, None)
        for col in range(0, sheet.max_column):
            cell = sheet.cell(row=row+1, column=col+1)
            if type(cell).__name__ != 'MergedCell':
                if cell.value == None or str(cell.value).strip() == '': # check if the cell is the last one with not None value
                    last = True
                    for colu in range(col+1, sheet.max_column):
                        if sheet.cell(row+1, colu+1).value != None and str(sheet.cell(row+1, colu+1).value).strip() != '':
                            last = False
                    if last:
                        break
                
                is_merged = False
                for mergedCell in merged_ranges:
                    if (cell.coordinate in mergedCell):
                        is_merged = True
                        start_col, start_row, end_row, end_col = mergedCell.bounds
                        rowspan = end_row - start_row + 1
                        colspan = end_col - start_col + 1
                        node = node.addkid(
                            WeirdNode(
                                set_node_label(conv(cell.value)), rowspan, colspan
                            )
                        )
                        merged_ranges.remove(mergedCell)
                        break
                if not is_merged:
                    node = node.addkid(
                        WeirdNode(
                            set_node_label(conv(cell.value)), 1, 1
                        )
                    )           
        tree = tree.addkid(node)
    return (tree)

def convert_json2tree(json_table, name, ocr_included=True):
    def compare_location(cell1, cell2):
        if cell1['rows'][0] == cell2['rows'][0]:
            return cell1['columns'][0] - cell2['columns'][0]
        else:
            return cell1['rows'][0] - cell2['rows'][0]

    def sort_cell_list(cell_list):
        return sorted(cell_list, key=cmp_to_key(compare_location))
    
    def set_node_label(label):
        if ocr_included:
            return label
        return name

    cell_list = sort_cell_list(json_table["cell_list"])
    conv = lambda i : i or ''
    tree = WeirdNode(set_node_label(name), None, None)
    row = 0
    node = WeirdNode(set_node_label('row_' + str(row+1)), None, None)
    for idx, lay in enumerate(cell_list):
        if lay['rows'][0] != row:
            tree = tree.addkid(node)
            row = lay['rows'][0]
            node = WeirdNode(set_node_label('row_' + str(row+1)), None, None)

        start_row, end_row, start_col, end_col = lay['rows'] + lay['columns']
        rowspan = end_row - start_row + 1
        colspan = end_col - start_col + 1
        node = node.addkid(
            WeirdNode(
                set_node_label(conv(combine_text_list(lay['text_list']))), rowspan, colspan
            )
        )
    return (tree)

def get_num_nodes(tree):
    t = len(WeirdNode.get_children(tree))
    t += sum([len(WeirdNode.get_children(node)) for node in WeirdNode.get_children(tree) ])
    return t

def cal_edit_distance(Ta, Tb):
    '''
    Ta, Tb: trees
    '''
    return distance(
        Ta, 
        Tb,
        get_children=WeirdNode.get_children,
        insert_cost=WeirdNode.insert_cost,
        remove_cost=WeirdNode.remove_cost,
        update_cost=WeirdNode.update_cost
    )

def cal_TEDS(Ta, Tb):
    '''
    Ta, Tb: trees
    '''
    denominator = 1 + max(get_num_nodes(Ta), get_num_nodes(Tb))
    return 1. - (cal_edit_distance(Ta, Tb)/float(denominator))

def compare_tables(table1, table2, name=None, ocr_included=True):
    if name:
        Ta = convert_json2tree(table1, name, ocr_included)
        Tb = convert_json2tree(table2, name, ocr_included)
    else:
        Ta = convert_sheet2tree(table1, ocr_included)
        Tb = convert_sheet2tree(table2, ocr_included)
    return cal_TEDS(Ta, Tb)

def compare_excels(gt_path, pd_path, ocr_included=True):
    apath = os.path.abspath(gt_path)
    gt_wb = load_workbook(apath)
    if not pd_path or not os.path.isfile(pd_path):
        return [0., len(gt_wb.worksheets), len(gt_wb.worksheets)*[0.]]

    bpath = os.path.abspath(pd_path)
    pd_wb = load_workbook(bpath)
    final = 0. # final is sum of all TEDS scores (not average) 
    results = []
    
    for idx, gt_sheet in enumerate(gt_wb.worksheets):
        name = gt_sheet.title
        if name in pd_wb.sheetnames:
            pd_sheet = pd_wb[name]
            res = compare_tables(gt_sheet, pd_sheet, ocr_included=ocr_included)
        else:
            res = 0.
        results.append(res)
        final += res
    return [final, len(gt_wb.worksheets), results]

def compare_jsons(gt_path, pd_path, ocr_included=True, threshold=0.8):
    """
    gt_path and pd_path are json files with spec io
    """
    apath = os.path.abspath(gt_path)
    gt_data = load_json(apath)
    gt_data = stand2spec_output(gt_data)
    if not pd_path or not os.path.isfile(pd_path):
        return [0., len(gt_data), len(gt_data)*[0.]]

    bpath = os.path.abspath(pd_path)
    pd_data = load_json(bpath)
    final = 0.
    results = []
    for idx, gt_table in enumerate(gt_data):
        name = 'table_' + str(idx+1)
        max_iou = 0.
        max_jdx = 0
        for jdx, pd_table in enumerate(pd_data):
            iou = cal_spec_location_iou(gt_table["location"], pd_table["location"])
            if iou > max_iou:
                max_iou = iou
                max_jdx = jdx
        if max_iou > threshold:
            pd_table = pd_data[max_jdx]
            res = compare_tables(gt_table, pd_table, name, ocr_included=ocr_included)
        else:
            res = 0.    
        results.append(res)
        final += res
    return [final, len(gt_data), results]

def compare(gt_dir, pd_dir, output_dir, extension='json', ocr_included=True):
    if extension == 'xlsx':
        cmp_func = compare_excels
    elif extension == 'json':
        cmp_func = compare_jsons
    else:
        assert 1, "Wrong extension ! Extension must be json or xlsx"

    gt_dir = os.path.abspath(gt_dir)
    filepaths = glob(gt_dir + '/*')
    TEDS_sum = 0
    TEDS_max = 0
    TEDS_min = 1
    mode = dict()
    n_sheets = 0
    summary = []
    
    for idx, filepath in enumerate(filepaths):
        filename = os.path.basename(filepath)
        pd_path = os.path.join(pd_dir, filename)
        final, num_tables, results = cmp_func(filepath, pd_path, ocr_included)        
        for res in results:
            mode[str(round(res,1))] = mode.get(str(round(res,1)), 0) + 1
            if TEDS_max < res:
                TEDS_max = res
            if TEDS_min > res:
                TEDS_min = res
        TEDS_sum += final
        n_sheets += num_tables
        file_analysis = {
            'file_name': filename,
            'number_of_table': num_tables,
            'teds_list': '    '.join([str(round(r, 4)) for r in results]),
            'mean_teds': round(final/float(num_tables), 4)
        }         
        summary.append(file_analysis)   
        
    visualize_bar_chart(mode, output_dir)
    dump_report_excel(summary, output_dir)

    TEDS_mode = float(max(mode, key=mode.get))
    TEDS_mean_per_gt = TEDS_sum/float(n_sheets)
    # TEDS_mean_per_pd = TEDS_sum/float(len(os.listdir(pd_dir)))

    TEDS_sum = round(TEDS_sum, 3)
    TEDS_max = round(TEDS_max, 3)
    TEDS_min = round(TEDS_min, 3)
    TEDS_mean_per_gt = round(TEDS_mean_per_gt, 3)
    # TEDS_mean_per_pd = round(TEDS_mean_per_pd, 3)
    return [TEDS_sum, TEDS_max, TEDS_min, TEDS_mode, TEDS_mean_per_gt]