import json
import cv2
import argparse
import os

# import xlrd
# import xlwt
# from xlwt import Workbook 
from openpyxl import load_workbook, Workbook

from utility import visualize_layout, visualize_table, check_layout_in_cell, locate_layouts, location2bbox

from table import Tee
from table.classes import Table
from layout import InvoiceLayout, JeffLayout
from ocr import CannetOCR

class Converter:
    def __init__(self, config):
        self.table_model = Tee(weights_path=config["table_weight"])
        self.layout_model = JeffLayout(config["layout_weight"])
        self.ocr_model = CannetOCR(config["ocr_weight"])
        self.OUTPUT_DIR = config["output_dir"]
        if not os.path.exists(self.OUTPUT_DIR):
            os.makedirs(self.OUTPUT_DIR)
        if not os.path.exists(os.path.join(self.OUTPUT_DIR, "visualization")):
            os.makedirs(os.path.join(self.OUTPUT_DIR, "visualization"))
        if not os.path.exists(os.path.join(self.OUTPUT_DIR, "prediction")):
            os.makedirs(os.path.join(self.OUTPUT_DIR, "prediction"))

    def visualize(self, extraction, origin_image, name, prefix=''):
        overview = visualize_layout(extraction['layout_output'], origin_image)
        overview = visualize_table(extraction['refined_boxes'], overview)
        cv2.imwrite(os.path.join(self.OUTPUT_DIR, "visualization", name, prefix + '-overview-' + name + '.png'), overview)
    
        for idx, (img, cells, layouts) in enumerate(zip(extraction['tables'], extraction['cells'], extraction['layouts'])):
            img = visualize_table(cells, img)
            img = visualize_layout(layouts, img)
            filename = prefix + 'table_' + str(idx+1) + '-' + name + '.png'
            path = os.path.join(self.OUTPUT_DIR, "visualization", name, filename)
            cv2.imwrite(path, img)

    def extract_coordinate(self, image, topdown=True):
        #pp2
        refined_boxes = self.table_model.process(image, resize=True, refine=True)
        result = self.layout_model.process(image)
    
        table_images = []
        layout_results = []
        cell_results = []
        table_locations = []

        sorted_tables = list(filter(lambda x: x["type"] == "table", refined_boxes))
        if topdown:
            sorted_tables.sort(key=lambda x: x['location'][0][1]) # this is a simple sort that needs to be updated 
    
        for entity in sorted_tables:
            if entity["type"] == "table":
                location = entity['location']
                table_locations.append(location)
                tlbr_poses = location2bbox(location)
                table = image[
                    tlbr_poses[1]:tlbr_poses[3], tlbr_poses[0]:tlbr_poses[2]
                ] # Image of table
                table_images.append(table)
                tab = {
                    'location': [tuple(tlbr_poses[:2]), (None, None), tuple(tlbr_poses[2:]), (None, None)]
                }
            
                res = [] # layouts that belongs to this table
                for idx, lay in enumerate(result):
                    if check_layout_in_cell(lay, tab):
                        tmp = dict()
                        tmp['location'] = [(x - tlbr_poses[0], y - tlbr_poses[1]) for (x,y) in lay['location']]
                        res.append(tmp) 
                layout_results.append(res)
            
                cells = []
                for idx in entity["contains"]:
                    cell = refined_boxes[idx].copy()
                    location = cell['location']
                    cell['location'] = [(x - tlbr_poses[0], y - tlbr_poses[1]) for (x,y) in location]
                    cells.append(cell)
                cell_results.append(cells)         
        return {
            'tables': table_images,
            'layouts': layout_results,
            'cells': cell_results,
            'table_locations': table_locations,
            'refined_boxes': refined_boxes,
            'layout_output': result
        }

    def extract_xlsx(self, table, layout_output, filepath, idx):
        if os.path.isfile(filepath):
            wb = load_workbook(filepath)
            wb.create_sheet()
            wb.active = idx
        else:
            wb = Workbook()
    
        sheet = wb.active 
        sheet.title = 'table_' + str(idx+1)
        for idx, lay in enumerate(layout_output):  
            x0, y0, x1, y1 = lay['bbox']
            if table[y0:y1, x0:x1].size == 0:
                field = {
                    'text': ''
                }
            else:
                field = self.ocr_model.process(table[y0:y1, x0:x1])
            
            if lay['line'] == 0:
                sheet.merge_cells(start_row=1, start_column=lay['col'][0] + 1, end_row=1, end_column=lay['col'][1] + 1) 
                sheet.cell(row = 1, column = lay['col'][0] + 1).value = field['text']
            else:
                sheet.cell(row = lay['line'] + 1, column = lay['col'] + 1).value = field['text']
        wb.save(filepath) 

    def convert(self, extraction, name , prefix=''):
        for idx, (img, cells, layouts) in enumerate(zip(extraction['tables'], extraction['cells'], extraction['layouts'])):
            try:
                layout_output = locate_layouts(cells, layouts)                
            except:
                continue
            filename = prefix + name + '.xlsx'
            path = os.path.join(self.OUTPUT_DIR, "prediction", filename)
            self.extract_xlsx(img, layout_output, path, idx)

    def run(self, path):
        name = os.path.basename(os.path.splitext(path)[0])
        if not os.path.exists(os.path.join(self.OUTPUT_DIR, "visualization", name)):
            os.makedirs(os.path.join(self.OUTPUT_DIR, "visualization", name))
        
        origin_image = cv2.imread(path)
        extraction = self.extract_coordinate(origin_image)
        self.visualize(extraction, origin_image, name)
        self.convert(extraction, name)

if __name__ == "__main__":

    # data = [
    #     '/Users/macbook/Documents/Cinnamon/tabxd/dataset/data_01/2018030500089_558404_04_2018030530006000730.tif',
    #     '/Users/macbook/Documents/Cinnamon/tabxd/dataset/data_01/2018030500008_558404_01_2018030530002000290.tif',
    #     '/Users/macbook/Documents/Cinnamon/tabxd/dataset/data_01/2018030500010_558404_01_2018030530002000390.tif',
    #     '/Users/macbook/Documents/Cinnamon/tabxd/dataset/data_01/2018030500011_558404_02_2018030530002000460.tif',
    #     '/Users/macbook/Documents/Cinnamon/tabxd/dataset/data_01/2018030500026_558404_02_2018030530003000380.tif',
    #     '/Users/macbook/Documents/Cinnamon/tabxd/dataset/data_01/2018030500025_558404_01_2018030530003000340.tif',
    #     '/Users/macbook/Documents/Cinnamon/tabxd/dataset/data_01/2018030500065_558404_02_2018030530005000440.tif',
    #     '/Users/macbook/Documents/Cinnamon/tabxd/dataset/data_01/2018030500084_558404_01_2018030530006000450.tif',
    #     '/Users/macbook/Documents/Cinnamon/tabxd/dataset/data_01/2018030500089_558404_05_2018030530006000740.tif',
    #     '/Users/macbook/Documents/Cinnamon/tabxd/dataset/data_01/2018030500107_558404_03_2018030530007000620.tif',
    #     '/Users/macbook/Documents/Cinnamon/tabxd/dataset/data_01/2018030500026_558404_01_2018030530003000370.tif',
    #     '/Users/macbook/Documents/Cinnamon/tabxd/dataset/data_01/2018030500214_558404_01_2018030530013000400.tif',
    #     '/Users/macbook/Documents/Cinnamon/tabxd/dataset/data_01/2018030500210_558404_02_2018030530013000290.tif',
    #     '/Users/macbook/Documents/Cinnamon/tabxd/dataset/data_01/2018030500210_558404_01_2018030530013000280.tif',
    #     '/Users/macbook/Documents/Cinnamon/tabxd/dataset/data_01/2018030500209_558404_01_2018030530013000220.tif',
    #     '/Users/macbook/Documents/Cinnamon/tabxd/dataset/data_01/2018030500208_558404_01_2018030530013000180.tif',
    #     '/Users/macbook/Documents/Cinnamon/tabxd/dataset/data_01/2018030500199_558404_01_2018030530012000500.tif',
    #     '/Users/macbook/Documents/Cinnamon/tabxd/dataset/data_01/2018030500184_558404_01_2018030530011000680.tif',
    #     '/Users/macbook/Documents/Cinnamon/tabxd/dataset/data_01/2018030500181_558404_01_2018030530011000580.tif',
    #     '/Users/macbook/Documents/Cinnamon/tabxd/dataset/data_01/2018030500178_558404_01_2018030530011000490.tif',
    #     '/Users/macbook/Documents/Cinnamon/tabxd/dataset/data_01/2018030500169_558404_01_2018030530011000220.tif',
    #     '/Users/macbook/Documents/Cinnamon/tabxd/dataset/data_01/2018030500166_558404_03_2018030530010000740.tif'
    # ]

    parser = argparse.ArgumentParser()
    parser.add_argument('-c', '--config_path', default='config.json', help='path to config')
    args = parser.parse_args()
    config = json.load(open(args.config_path))
    converter = Converter(config)

    DATA_DIR = config["data_dir"]
    data = os.listdir(DATA_DIR)

    for t in data:
        try:
            converter.run(os.path.join(DATA_DIR, t))
        except:
            print(t)




